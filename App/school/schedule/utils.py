from school.models import ChromeBrowser
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from dataclasses import dataclass
from datetime import datetime
import pandas as pd
import os
import re
import openpyxl


@dataclass
class Subject:
    '''Class to represent a subject'''
    name: str
    teacher: str
    classroom: str
    day: str
    startTime: datetime
    endTime: datetime
    startdate: datetime
    enddate: datetime
    group: str

    def __str__(self) -> str:
        return f'Subject:{" - ".join([f"{column.name}:{getattr(self, column.name)}" for column in self.__table__.columns])})'

    def to_dict(self) -> dict:
        '''Convert the subject to a dictionary'''
        return self.__dict__


def findScheduleTable(browser):
    try:
        scheduleContent = browser.find_element(By.ID, "contenido-tabla")
        print("Schedule content found ✅")
    except NoSuchElementException:
        print("Schedule content not found ❌")
    return scheduleContent


def findScheduleSubjects(scheduleContent: str) -> list[str]:
    '''Extracts the schedule subjects from the schedule content'''
    try:
        rows = scheduleContent.find_elements(By.CSS_SELECTOR, "div.row")
        print(f"Schedule content has {len(rows)} rows🔎")
    except NoSuchElementException:
        print("Schedule content has no rows ❌")

    data = []
    for row in rows:
        # Find all the div elements within the row
        cells = row.find_elements(By.CSS_SELECTOR, 'div')

        # Extract the data from the cells
        cell_data = [cell.text for cell in cells]

        # Add the data to the list
        data.append(cell_data)

    return data


def splitScheduleSubjects(scheduleRows: list[list[str]]) -> list[Subject]:
    '''Splits the schedule subjects into a list of subjects'''

    return loadScheduleData(scheduleRows, current_day='')


def cleanScheduleData(data: list[str]) -> dict[str, str]:
    '''Cleans the schedule data'''
    day = data[1].strip()
    start_time = data[2].strip()
    end_time = data[3].strip()
    subject = data[4].strip()
    teacher = data[6].strip()
    start_date = data[7].strip()
    end_date = data[8].strip()
    group = data[9].strip()
    classroom = re.compile(
        r'([^/]*)$').search(re.sub(r'\n', '', data[5].strip())).group(1).replace('Ver', '').lstrip()

    return {
        'day': day,
        'start_time': start_time,
        'end_time': end_time,
        'subject': subject,
        'teacher': teacher,
        'start_date': start_date,
        'end_date': end_date,
        'group': group,
        'classroom': classroom
    }


def loadScheduleData(scheduleRows: list[list[str]]) -> list[Subject]:
    '''Loads the schedule data into a Subject object'''
    objects: list[Subject] = []
    current_day = ''
    for data in scheduleRows:
        subject_data = cleanScheduleData(data)
        subject_data['day'] = subject_data['day'] or current_day
        objects.append(createSubject(**subject_data))
        current_day = subject_data['day'] if subject_data['day'] else current_day

    return objects


def createSubject(day: str, start_time: datetime, end_time: datetime, subject: str, teacher: str, start_date: datetime, end_date: datetime, group: str, classroom: str):
    '''Creates a subject object'''

    # convert start and end time to datetime
    subject = Subject(day=day, startTime=datetime.strptime(end_time, '%H:%M'), endTime=datetime.strptime(start_time, '%H:%M'), name=subject, teacher=teacher,
                      startdate=start_date, enddate=end_date, group=group, classroom=classroom)
    return subject


def getScheduleContent(browser: ChromeBrowser) -> list[list[str]]:
    '''Extracts the schedule content from the schedule page'''
    try:
        content = loadScheduleData(
            findScheduleSubjects(findScheduleTable(browser)))
        print("Schedule content extracted ✅")
    except Exception as e:
        print(f"Schedule content not extracted ❌: {e}")
        content = []
    return content


def formatSubject(subjects: list[Subject]) -> str:
    '''Formats the subjects'''
    name_length = max(len(subject.name) for subject in subjects)
    teacher_length = max(len(subject.teacher) for subject in subjects)
    classroom_length = max(len(subject.classroom) for subject in subjects)
    day_length = max(len(subject.day) for subject in subjects)
    start_time_length = max(len(subject.startTime) for subject in subjects)
    end_time_length = max(len(subject.endTime) for subject in subjects)
    start_date_length = max(len(subject.startdate) for subject in subjects)
    end_date_length = max(len(subject.enddate) for subject in subjects)
    group_length = max(len(subject.group) for subject in subjects)
    # Print the header
    print(f"{'Name':<{name_length}}  {'Teacher':<{teacher_length}}  {'Classroom':<{classroom_length}}  {'Day':<{day_length}}  {'Start Time':<{start_time_length}}  {'End Time':<{end_time_length}}  {'Start Date':<{start_date_length}}  {'End Date':<{end_date_length}}  {'Group':<{group_length}}")
    print("-" * (name_length + teacher_length + classroom_length + day_length +
          start_time_length + end_time_length + start_date_length + end_date_length + group_length + 8))
    # Print the rows
    for subject in subjects:
        print(f"{subject.name:<{name_length}}  {subject.teacher:<{teacher_length}}  {subject.classroom:<{classroom_length}}  {subject.day:<{day_length}}  {subject.startTime:<{start_time_length}}  {subject.endTime:<{end_time_length}}  {subject.startdate:<{start_date_length}}  {subject.enddate:<{end_date_length}}  {subject.group:<{group_length}}")


def scheduleExcel(subjects: list[Subject]) -> pd:
    '''Exports the schedule to an excel file'''
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()

    # Get the active worksheet
    worksheet = workbook.active

    # Set the column width
    worksheet.column_dimensions['A'].width = 10
    for i in range(2, 7):
        worksheet.column_dimensions[f"{chr(i+64)}"].width = 10

    # Leave cell A1 blank
    worksheet.cell(row=1, column=1).value = ""

    # Write the hours and half-hours in column A, starting at row 2
    for hour in range(7, 21):
        worksheet.cell(row=hour*2-11, column=1).value = f"{hour}:00"
        worksheet.cell(row=hour*2-10, column=1).value = f"{hour}:30"

    # Write the days of the week in Spanish in columns B to G, starting at row 1
    days_of_week_spanish = ["Lunes", "Martes", "Miércoles",
                            "Jueves", "Viernes", ]
    for i, day in enumerate(days_of_week_spanish):
        worksheet.cell(row=1, column=i+2).value = day

    #  Write subject content in the cells
    # for subject in subjects:
    #     start_time = subject.startTime
    #     end_time = subject.endTime
    #     day = subject.day
    #     # Get the row and column of the start time
    #     start_row = (start_time * 2) - 11
    #     start_column = days_of_week_spanish.index(day) + 2
    #     # Get the row and column of the end time
    #     end_row = (end_time * 2) - 10
    #     end_column = days_of_week_spanish.index(day) + 2
    #     # Write the subject name in the cells
    #     for row in range(start_row, end_row+1):
    #         for column in range(start_column, end_column+1):
    #             worksheet.cell(row=row, column=column).value = subject.name

    # Check if the file already exists and delete it if it does
    if os.path.exists("hours.xlsx"):
        os.remove("hours.xlsx")

    # Save the workbook
    workbook.save("hours.xlsx")

    print("Hours and days of the week written to Excel file successfully!")
# df = pd.DataFrame(subjects)
# df.to_excel('schedule.xlsx', index=False)
